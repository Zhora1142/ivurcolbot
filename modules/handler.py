from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from os import remove


class Handler:
    def __init__(self):
        self.original_wb = load_workbook('files/file.xlsx', data_only=True)
        self.new_wb = Workbook()

    def table_coordinates(self):
        sheet = self.original_wb.active

        begin = {'row': None, 'column': 1}
        end = {'row': None, 'column': 5}

        for row in range(1, sheet.max_row + 1):

            if sheet.cell(row=row, column=1).value == 'Время':
                begin['row'] = row + 2

            if begin['row']:
                if sheet.cell(row=row, column=1).value is None and sheet.cell(row=row + 1, column=1).value is None:
                    end['row'] = row - 1
                    break

                if row == sheet.max_row:
                    end['row'] = row

        if not begin['row'] or not end['row']:
            return None
        else:
            return {'begin': begin, 'end': end}

    def move_table(self):
        coordinates = self.table_coordinates()

        if coordinates:
            begin_cell = f'{get_column_letter(coordinates["begin"]["column"])}{coordinates["begin"]["row"]}'
            end_cell = f'{get_column_letter(coordinates["end"]["column"])}{coordinates["end"]["row"]}'

            old_sheet = self.original_wb.active
            new_sheet = self.new_wb.active

            cells = old_sheet[begin_cell:end_cell]

            for row in range(len(cells)):
                for column in range(5):
                    if type(cells[row][column].value) == str:
                        cells[row][column].value = cells[row][column].value.replace('\n', ' ')

                    new_sheet.cell(row=row + 1, column=column + 1).value = cells[row][column].value

            for row in range(1, new_sheet.max_row + 1):
                if len(new_sheet.cell(row=row, column=1).value) > 10:
                    new_sheet.cell(row=row, column=1).value = None

                if type(new_sheet.cell(row=row, column=4).value) == str:
                    new_value = new_sheet.cell(row=row, column=4).value.replace('Э', '3').replace('б', '6')

                    if new_sheet.cell(row=row, column=4).value != new_value:
                        new_sheet.cell(row=row, column=4).value = int(new_value)

            self.new_wb.save('files/result.xlsx')

            remove('files/file.jpg')
            remove('files/file.xlsx')
        else:
            raise Exception('Parsing_error')
