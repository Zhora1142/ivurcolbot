from openpyxl import load_workbook
from modules.msql_collector import MysqlCollector
import yaml
from json import dumps
from os import remove

with open('config.yml') as stream:
    config = yaml.safe_load(stream)['sql']


class FinalHandler:
    def __init__(self):
        self.wb = load_workbook('files/result.xlsx')
        self.collector = MysqlCollector(**config, db='ivurcol')

    def reset(self):
        for day in range(1, 7):
            for couple in range(1, 7):
                data = {'name': None, 'type': None, 'room': None, 'tutor': None}
                result = self.collector.update(table='timetable',
                                               values={f'couple_{couple}': dumps(data)},
                                               where=f'day={day}')
                if not result['status']:
                    raise Exception(result['data'])

    def handle(self):
        self.reset()

        sheet = self.wb.active

        if sheet.cell(row=1, column=1).value is None:
            raise(Exception('wrong file'))

        day, couple, flag = 1, 1, True
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value:
                while flag:
                    if str(couple) not in sheet.cell(row=row, column=1).value:
                        couple += 1
                    else:
                        flag = False

                if sheet.cell(row=row, column=2):
                    data = {
                        'name': sheet.cell(row=row, column=2).value,
                        'type': sheet.cell(row=row, column=3).value,
                        'room': sheet.cell(row=row, column=4).value,
                        'tutor': sheet.cell(row=row, column=5).value
                    }

                    result = self.collector.update(table='timetable',
                                                   values={f'couple_{couple}': dumps(data, ensure_ascii=False)},
                                                   where=f'day={day}')
                    if not result['status']:
                        raise Exception(result['data'])
                    couple += 1
            else:
                couple = 1
                day += 1
                flag = True

        remove('files/result.xlsx')
